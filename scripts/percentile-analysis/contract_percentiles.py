#!/usr/bin/env python3
"""
Contract-Level Percentile Performance
======================================
For every H+R contract and every measure, calculates the raw score
and percentile rank within the distribution of all H+R contracts.

Supports two percentile methods (configurable via --method flag):
  - percentrank_inc  : Excel's PERCENTRANK.INC — (count below) / (n-1) * 100 [default]
  - percentileofscore: scipy — (count at or below) / n * 100

Usage:
  python contract_percentiles.py --data-dir /path/to/csvs --output results.xlsx
  python contract_percentiles.py --data-dir /path/to/csvs --output results.xlsx --method percentileofscore
  python contract_percentiles.py --data-dir /path/to/csvs --output results.json --format json
"""

import argparse
import json
import os
import re
import sys

import numpy as np
import pandas as pd

from config import (
    FILES, CONTRACT_PREFIXES, ALL_METHODS, is_inverted, normalize, calc_percentile,
)


def parse_args():
    parser = argparse.ArgumentParser(description="Calculate contract-level percentile performance")
    parser.add_argument("--data-dir", default=os.path.join(os.path.dirname(__file__), "data"),
                        help="Directory containing the CSV measure data files (default: ./data)")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx or .json)")
    parser.add_argument("--format", choices=["xlsx", "json"], default=None, help="Output format (auto-detected from extension if omitted)")
    parser.add_argument("--method", choices=ALL_METHODS, default="percentrank_inc",
                        help="Percentile calculation method (default: percentrank_inc)")
    parser.add_argument("--years", nargs="+", type=int, default=None, help="Specific years to process (default: all)")
    return parser.parse_args()


def load_and_calculate(data_dir: str, method: str, years: list[int] | None = None) -> dict:
    """Load CSVs, filter to H+R contracts, calculate percentile ranks.

    Returns dict: {year: {"df": DataFrame, "measures": [(code, display), ...]}}
    """
    all_data = {}

    for yr, (filename, skip_rows, code_row) in FILES.items():
        if years and yr not in years:
            continue

        filepath = os.path.join(data_dir, filename)
        if not os.path.exists(filepath):
            print(f"  WARNING: {filename} not found, skipping {yr}")
            continue

        raw = pd.read_csv(filepath, encoding="latin-1", header=None)

        # Filter to H+R contracts
        contracts = raw.iloc[skip_rows:, 0].astype(str).str.strip()
        hr_mask = contracts.str.startswith(CONTRACT_PREFIXES[0])
        for prefix in CONTRACT_PREFIXES[1:]:
            hr_mask = hr_mask | contracts.str.startswith(prefix)
        hr_indices = contracts[hr_mask].index

        contract_ids = raw.loc[hr_indices, 0].astype(str).str.strip().values
        contract_names = raw.loc[hr_indices, 2].astype(str).str.strip().values
        org_names = raw.loc[hr_indices, 3].astype(str).str.strip().values

        # Identify measure columns
        codes = raw.iloc[code_row]
        measure_cols = []
        for ci, val in codes.items():
            if isinstance(val, str) and val.strip() and ":" in val:
                display = re.sub(r"^[CD]\d+:\s*", "", val.strip())
                display = re.sub(r"[^\x20-\x7e]", " ", display).strip()
                display = re.sub(r"\s+", " ", display)
                code = val.strip().split(":")[0].strip()
                measure_cols.append((ci, code, display))

        # Build dataframe
        rows = [
            {"Contract_ID": contract_ids[i], "Contract_Name": contract_names[i], "Org_Name": org_names[i]}
            for i in range(len(hr_indices))
        ]
        df = pd.DataFrame(rows)

        # Calculate percentiles per measure
        for ci, code, display in measure_cols:
            raw_vals = raw.loc[hr_indices, ci].astype(str).str.strip().str.rstrip("%")
            numeric_vals = pd.to_numeric(raw_vals, errors="coerce")
            valid_vals = numeric_vals.dropna().values
            inv = is_inverted(display)

            df[f"{code}_Score"] = numeric_vals.values

            pcts = np.full(len(hr_indices), np.nan)
            if len(valid_vals) >= 2:
                for i, val in enumerate(numeric_vals.values):
                    if pd.notna(val):
                        pcts[i] = calc_percentile(valid_vals, val, inverted=inv, method=method)

            df[f"{code}_Pctile"] = np.round(pcts, 1)

        all_data[yr] = {
            "df": df,
            "measures": [(code, display) for ci, code, display in measure_cols],
        }

        h_count = sum(1 for c in contract_ids if c.startswith("H"))
        r_count = sum(1 for c in contract_ids if c.startswith("R"))
        print(f"  {yr}: {len(df)} contracts (H={h_count}, R={r_count}), {len(measure_cols)} measures [{method}]")

    return all_data


def save_json(all_data: dict, output_path: str, method: str):
    """Save results as JSON."""
    result = {"method": method, "years": {}}

    for yr, info in sorted(all_data.items()):
        df = info["df"]
        measures = info["measures"]
        yr_data = []

        for _, row in df.iterrows():
            contract = {
                "contract_id": row["Contract_ID"],
                "contract_name": row["Contract_Name"],
                "org_name": row["Org_Name"],
                "measures": {},
            }
            for code, display in measures:
                score = row.get(f"{code}_Score")
                pctile = row.get(f"{code}_Pctile")
                if pd.notna(score):
                    contract["measures"][code] = {
                        "name": display,
                        "score": float(score),
                        "percentile": float(pctile) if pd.notna(pctile) else None,
                        "inverted": is_inverted(display),
                    }
            yr_data.append(contract)

        result["years"][str(yr)] = yr_data

    with open(output_path, "w") as f:
        json.dump(result, f, indent=2)
    print(f"\nSaved JSON to {output_path}")


def save_xlsx(all_data: dict, output_path: str, method: str):
    """Save results as formatted Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    measure_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_fill_score = PatternFill("solid", fgColor="E8E8E8")
    sub_fill_pctile = PatternFill("solid", fgColor="F5F5F5")
    sub_font = Font(bold=True, size=8, name="Arial")
    data_font = Font(size=9, name="Arial")
    id_font = Font(bold=True, size=9, name="Arial")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    def pctile_fill(val):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        if val >= 80:
            return PatternFill("solid", fgColor="C6EFCE")
        elif val >= 60:
            return PatternFill("solid", fgColor="D6E4F0")
        elif val >= 30:
            return PatternFill("solid", fgColor="FFEB9C")
        elif val >= 15:
            return PatternFill("solid", fgColor="FFC7CE")
        else:
            return PatternFill("solid", fgColor="FF6B6B")

    method_label = "PERCENTRANK.INC" if method == "percentrank_inc" else "percentileofscore"

    for yr in sorted(all_data.keys()):
        df = all_data[yr]["df"]
        measures = all_data[yr]["measures"]
        ws = wb.create_sheet(title=str(yr))

        total_cols = 3 + len(measures) * 2
        end_col = get_column_letter(total_cols)
        ws.merge_cells(f"A1:{end_col}1")
        ws["A1"] = f"Contract-Level Percentile Performance (H+R Contracts) — {yr} Star Ratings"
        ws["A1"].font = Font(bold=True, size=13, name="Arial", color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"A2:{end_col}2")
        ws["A2"] = f"{len(df)} H+R contracts × {len(measures)} measures | Method: {method_label} | Score = raw value | %ile = percentile rank"
        ws["A2"].font = Font(italic=True, size=9, name="Arial", color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        for ci, label in enumerate(["Contract ID", "Contract Name", "Organization"], 1):
            ws.merge_cells(start_row=4, start_column=ci, end_row=5, end_column=ci)
            c = ws.cell(row=4, column=ci, value=label)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border

        for mi, (code, display) in enumerate(measures):
            col_start = 4 + mi * 2
            col_end = col_start + 1
            inv = is_inverted(display)
            label = f"{code}: {display}" + (" ↓" if inv else "")

            ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)
            c = ws.cell(row=4, column=col_start, value=label)
            c.font = Font(bold=True, size=8, name="Arial", color="8B0000" if inv else "1F4E79")
            c.fill = measure_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = thin_border

            c_s = ws.cell(row=5, column=col_start, value="Score")
            c_s.font = sub_font; c_s.fill = sub_fill_score; c_s.alignment = Alignment(horizontal="center"); c_s.border = thin_border
            c_p = ws.cell(row=5, column=col_end, value="%ile")
            c_p.font = sub_font; c_p.fill = sub_fill_pctile; c_p.alignment = Alignment(horizontal="center"); c_p.border = thin_border

        for ri, (_, row) in enumerate(df.iterrows()):
            data_row = 6 + ri
            ws.cell(row=data_row, column=1, value=row["Contract_ID"]).font = id_font
            ws.cell(row=data_row, column=1).border = thin_border
            ws.cell(row=data_row, column=2, value=row["Contract_Name"]).font = data_font
            ws.cell(row=data_row, column=2).border = thin_border
            ws.cell(row=data_row, column=3, value=row["Org_Name"]).font = data_font
            ws.cell(row=data_row, column=3).border = thin_border

            for mi, (code, display) in enumerate(measures):
                col_score = 4 + mi * 2
                col_pctile = col_score + 1
                score_val = row.get(f"{code}_Score")
                pctile_val = row.get(f"{code}_Pctile")

                c_s = ws.cell(row=data_row, column=col_score)
                if score_val is not None and not (isinstance(score_val, float) and np.isnan(score_val)):
                    c_s.value = score_val
                    c_s.number_format = "0.00" if isinstance(score_val, float) and score_val < 10 else "0"
                    c_s.font = data_font
                else:
                    c_s.value = "—"
                    c_s.font = Font(size=9, name="Arial", color="BBBBBB")
                c_s.alignment = Alignment(horizontal="center")
                c_s.border = thin_border

                c_p = ws.cell(row=data_row, column=col_pctile)
                if pctile_val is not None and not (isinstance(pctile_val, float) and np.isnan(pctile_val)):
                    c_p.value = pctile_val
                    c_p.number_format = "0.0"
                    fill = pctile_fill(pctile_val)
                    if fill:
                        c_p.fill = fill
                    c_p.font = data_font
                else:
                    c_p.value = "—"
                    c_p.font = Font(size=9, name="Arial", color="BBBBBB")
                c_p.alignment = Alignment(horizontal="center")
                c_p.border = thin_border

        ws.column_dimensions["A"].width = 11
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25
        for mi in range(len(measures)):
            ws.column_dimensions[get_column_letter(4 + mi * 2)].width = 7
            ws.column_dimensions[get_column_letter(5 + mi * 2)].width = 7
        ws.freeze_panes = "D6"

        legend_row = 6 + len(df) + 2
        ws.cell(row=legend_row, column=1, value="Percentile Color Legend:").font = Font(bold=True, size=9, name="Arial")
        for j, (lbl, color) in enumerate([("≥80th (5★)", "C6EFCE"), ("60-79th (4★)", "D6E4F0"),
                                           ("30-59th (3★)", "FFEB9C"), ("15-29th (2★)", "FFC7CE"), ("<15th (1★)", "FF6B6B")]):
            c = ws.cell(row=legend_row, column=3 + j * 2, value=lbl)
            c.font = Font(size=8, name="Arial")
            c.fill = PatternFill("solid", fgColor=color)
        ws.cell(row=legend_row + 1, column=1,
                value="↓ = Inverted measure (lower score = better). Percentiles flipped so higher %ile = better.").font = Font(italic=True, size=8, name="Arial", color="8B0000")

    wb.save(output_path)
    print(f"\nSaved Excel to {output_path}")


def main():
    args = parse_args()
    fmt = args.format or ("json" if args.output.endswith(".json") else "xlsx")

    print(f"Contract Percentile Performance — method: {args.method}")
    print(f"Loading data from: {args.data_dir}\n")

    all_data = load_and_calculate(args.data_dir, args.method, args.years)

    if not all_data:
        print("ERROR: No data loaded. Check --data-dir path and CSV filenames.")
        sys.exit(1)

    if fmt == "json":
        save_json(all_data, args.output, args.method)
    else:
        save_xlsx(all_data, args.output, args.method)


if __name__ == "__main__":
    main()
