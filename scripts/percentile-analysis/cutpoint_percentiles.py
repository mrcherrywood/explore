#!/usr/bin/env python3
"""
Cut Point Percentile Equivalents
==================================
For each measure and year, finds what percentile each CMS cut point
corresponds to in the actual H+R contract data distribution.

Supports two percentile methods (configurable via --method flag):
  - percentrank_inc  : Excel's PERCENTRANK.INC — (count below) / (n-1) * 100 [default]
  - percentileofscore: scipy — (count at or below) / n * 100

Usage:
  python cutpoint_percentiles.py --data-dir /path/to/csvs --cut-points /path/to/cutpoints.xlsx --output results.xlsx
  python cutpoint_percentiles.py --data-dir /path/to/csvs --cut-points /path/to/cutpoints.xlsx --output results.json --method percentileofscore
"""

import argparse
import json
import os
import re
import sys

import numpy as np
import pandas as pd

from config import (
    FILES, CONTRACT_PREFIXES, MANUAL_CP_TO_NORM, is_inverted, normalize, calc_percentile,
)


def parse_args():
    parser = argparse.ArgumentParser(description="Calculate cut point percentile equivalents")
    parser.add_argument("--data-dir", default=os.path.join(os.path.dirname(__file__), "data"),
                        help="Directory containing the CSV measure data files (default: ./data)")
    parser.add_argument("--cut-points",
                        default=os.path.join(os.path.dirname(__file__), "data", "Stars 2016-2028 Cut Points 12.2025 (1).xlsx"),
                        help="Path to the cut points Excel file (default: ./data/Stars...xlsx)")
    parser.add_argument("--output", required=True, help="Output file path (.xlsx or .json)")
    parser.add_argument("--format", choices=["xlsx", "json"], default=None, help="Output format (auto-detected if omitted)")
    parser.add_argument("--method", choices=["percentrank_inc", "percentileofscore"], default="percentrank_inc",
                        help="Percentile calculation method (default: percentrank_inc)")
    parser.add_argument("--years", nargs="+", type=int, default=None, help="Specific years to process (default: all)")
    return parser.parse_args()


def find_match(cp_name, cp_code_prefix, col_names_norm, col_indices):
    """Find the matching column in raw data for a cut point measure name."""
    mn = normalize(cp_name)
    mapped = MANUAL_CP_TO_NORM.get(cp_name, "").lower()

    for idx, nv in zip(col_indices, col_names_norm):
        if mn and (mn in nv or nv in mn):
            if "call center" in mn:
                if cp_code_prefix == "C" and ("partd" in mn or "part d" in mn):
                    continue
                if cp_code_prefix == "D" and ("partc" in mn or "part c" in mn):
                    continue
            if "members choosing" in mn:
                if "partd" in nv or "part d" in nv:
                    continue
            return idx
        if mapped and (mapped in nv or nv in mapped):
            if "call center" in mapped:
                if cp_code_prefix == "C" and ("partd" in mapped or "part d" in mapped):
                    continue
                if cp_code_prefix == "D" and ("partc" in mapped or "part c" in mapped):
                    continue
            return idx
    return None


def get_dist_context(vals: np.ndarray, inv: bool) -> dict:
    """Compute distribution statistics and generate context notes."""
    p25, p50, p75 = [float(x) for x in np.percentile(vals, [25, 50, 75])]
    iqr = p75 - p25
    skew = float(pd.Series(vals).skew())

    notes = []
    if inv:
        if p25 < 1 and p75 < 5:
            notes.append(f"Most plans score very low (median {p50:.1f})")
        if skew > 1.5:
            notes.append("heavy right tail (few plans with very high/poor scores)")
    else:
        if p75 > 95 and (float(vals.max()) - p75) < 3:
            notes.append(f"ceiling effect — {(vals >= 95).sum()}/{len(vals)} plans score ≥95")
        elif p75 > 90 and iqr < 5:
            notes.append(f"compressed at top — IQR only {iqr:.0f} pts, median {p50:.0f}")
        if iqr < 3:
            notes.append(f"very narrow spread (IQR={iqr:.1f})")
        elif iqr < 8:
            notes.append(f"narrow spread (IQR={iqr:.1f})")
        if skew < -1.5:
            notes.append("strong left skew (long low-performance tail)")
        elif skew < -0.8:
            notes.append("moderately left-skewed")
        elif skew > 1.5:
            notes.append("strong right skew")
        elif skew > 0.8:
            notes.append("moderately right-skewed")
    if not inv and float(vals.max()) - float(vals.min()) < 15:
        notes.append(f"tight range ({vals.min():.0f}–{vals.max():.0f})")
    if not notes:
        if iqr > 20:
            notes.append(f"wide spread (IQR={iqr:.0f})")
        else:
            notes.append(f"moderate spread (IQR={iqr:.0f})")

    return {
        "context": "; ".join(notes),
        "median": float(p50),
        "iqr": float(iqr),
        "p25": float(p25),
        "p75": float(p75),
        "min": float(vals.min()),
        "max": float(vals.max()),
        "skew": round(skew, 2),
        "range": f"{vals.min():.0f}–{vals.max():.0f}",
    }


def load_and_calculate(data_dir: str, cut_points_path: str, method: str, years: list[int] | None = None) -> tuple:
    """Load data, match cut points to measures, calculate percentile equivalents.

    Returns (all_results, all_dist_context) dicts.
    """
    cp_df = pd.read_excel(cut_points_path, sheet_name="Cut Points")
    cp_df = cp_df.rename(columns=lambda c: c.strip())

    all_results = {}
    all_dist_context = {}

    for yr, (filename, skip_rows, code_row) in FILES.items():
        yr_int = yr
        if years and yr not in years:
            continue

        filepath = os.path.join(data_dir, filename)
        if not os.path.exists(filepath):
            print(f"  WARNING: {filename} not found, skipping {yr}")
            continue

        raw = pd.read_csv(filepath, encoding="latin-1", header=None)

        contracts = raw.iloc[skip_rows:, 0].astype(str).str.strip()
        hr_mask = contracts.str.startswith(CONTRACT_PREFIXES[0])
        for prefix in CONTRACT_PREFIXES[1:]:
            hr_mask = hr_mask | contracts.str.startswith(prefix)
        h_indices = contracts[hr_mask].index

        codes = raw.iloc[code_row]
        col_names_norm, col_indices = [], []
        for ci, val in codes.items():
            if isinstance(val, str) and val.strip():
                nv = normalize(val.strip())
                if nv:
                    col_names_norm.append(nv)
                    col_indices.append(ci)

        yr_cp = cp_df[cp_df["StarsYear"] == yr_int]
        seen_cp = set()
        yr_results = []

        for _, row in yr_cp.iterrows():
            cp_name = str(row["MeasureName"]).strip()
            hl_code = str(row.get("HLCode", "")).strip()
            cp_prefix = hl_code[0] if hl_code else ""
            if cp_name in seen_cp:
                continue
            seen_cp.add(cp_name)

            try:
                cp2 = float(row["2Star"])
                cp3 = float(row["3Star"])
                cp4 = float(row["4Star"])
                cp5 = float(row["5Star"])
            except (ValueError, TypeError):
                continue

            col_idx = find_match(cp_name, cp_prefix, col_names_norm, col_indices)
            if col_idx is None:
                continue

            series = raw.loc[h_indices, col_idx].astype(str).str.strip().str.rstrip("%")
            vals = pd.to_numeric(series, errors="coerce").dropna().values
            if len(vals) < 5:
                continue

            inv = is_inverted(cp_name)
            rec = {"measure": cp_name, "n": len(vals)}

            for star, cp_val in [("2", cp2), ("3", cp3), ("4", cp4), ("5", cp5)]:
                pct = calc_percentile(vals, cp_val, inverted=inv, method=method)
                rec[f"cp{star}"] = cp_val
                rec[f"cp{star}_pct"] = pct

            yr_results.append(rec)
            all_dist_context[f"{yr}|{cp_name}"] = get_dist_context(vals, inv)

        all_results[str(yr)] = yr_results
        print(f"  {yr}: {len(yr_results)} measures matched [{method}]")

    return all_results, all_dist_context


def save_json(all_results: dict, all_dist_context: dict, output_path: str, method: str):
    """Save results as JSON."""
    output = {
        "method": method,
        "results": all_results,
        "distributions": all_dist_context,
    }
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)
    print(f"\nSaved JSON to {output_path}")


def save_xlsx(all_results: dict, all_dist_context: dict, output_path: str, method: str):
    """Save results as formatted Excel workbook with per-year tabs, cross-year comparison, and summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    header_font_white = Font(bold=True, size=11, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    data_font = Font(size=10, name="Arial")
    context_font = Font(size=9, name="Arial", color="555555", italic=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"), right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"), bottom=Side(style="thin", color="B0B0B0"),
    )

    expected = {"2★": 15, "3★": 30, "4★": 60, "5★": 80}
    legends = [("≤ 5 pp", "C6EFCE"), ("5–15 pp", "FFEB9C"), ("15–25 pp", "FFC7CE"), ("> 25 pp", "FF6B6B")]

    def delta_fill(delta_abs):
        if delta_abs <= 5:
            return PatternFill("solid", fgColor="C6EFCE")
        elif delta_abs <= 15:
            return PatternFill("solid", fgColor="FFEB9C")
        elif delta_abs <= 25:
            return PatternFill("solid", fgColor="FFC7CE")
        else:
            return PatternFill("solid", fgColor="FF6B6B")

    method_label = "PERCENTRANK.INC" if method == "percentrank_inc" else "percentileofscore"
    years = sorted(all_results.keys())

    # --- Per-year tabs ---
    for yr in years:
        ws = wb.create_sheet(title=yr)
        measures = all_results[yr]

        ws.merge_cells("A1:N1")
        ws["A1"] = f"Cut Point Percentile Equivalents (H+R Contracts) — {yr} Star Ratings [{method_label}]"
        ws["A1"].font = Font(bold=True, size=14, name="Arial", color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Measure", "N", "2★ Cut Pt", "2★ Actual %ile", "3★ Cut Pt", "3★ Actual %ile",
                   "4★ Cut Pt", "4★ Actual %ile", "5★ Cut Pt", "5★ Actual %ile",
                   "Median", "IQR", "Range", "Distribution Notes"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = header_font_white
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = thin_border

        for i, m in enumerate(measures):
            row = 4 + i
            inv = is_inverted(m["measure"])
            label = m["measure"] + (" ↓" if inv else "")
            ws.cell(row=row, column=1, value=label).font = Font(size=10, name="Arial", italic=inv, color="8B0000" if inv else "000000")
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=m["n"]).font = data_font
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2).border = thin_border

            for si, (sk, ev) in enumerate([("2", 15), ("3", 30), ("4", 60), ("5", 80)]):
                cp_col = 3 + si * 2
                pct_col = 4 + si * 2
                cp_val = m[f"cp{sk}"]
                pct_val = m[f"cp{sk}_pct"]

                c_cp = ws.cell(row=row, column=cp_col, value=cp_val)
                c_cp.font = data_font
                c_cp.alignment = Alignment(horizontal="center")
                c_cp.border = thin_border
                c_cp.number_format = "0.00" if isinstance(cp_val, float) and cp_val < 10 else "0"

                c_pct = ws.cell(row=row, column=pct_col, value=pct_val)
                c_pct.font = data_font
                c_pct.alignment = Alignment(horizontal="center")
                c_pct.border = thin_border
                c_pct.number_format = "0.0"
                if isinstance(pct_val, (int, float)):
                    c_pct.fill = delta_fill(abs(pct_val - ev))

            ctx = all_dist_context.get(f"{yr}|{m['measure']}", {})
            for col_idx, key in [(11, "median"), (12, "iqr")]:
                val = ctx.get(key)
                c = ws.cell(row=row, column=col_idx, value=round(val, 1) if val is not None else "—")
                c.font = data_font
                c.alignment = Alignment(horizontal="center")
                c.border = thin_border
            ws.cell(row=row, column=13, value=ctx.get("range", "—")).font = data_font
            ws.cell(row=row, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=13).border = thin_border
            ws.cell(row=row, column=14, value=ctx.get("context", "")).font = context_font
            ws.cell(row=row, column=14).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=14).border = thin_border

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 7
        for col in range(3, 11):
            ws.column_dimensions[get_column_letter(col)].width = 13
        ws.column_dimensions["K"].width = 9
        ws.column_dimensions["L"].width = 8
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 45

    wb.save(output_path)
    print(f"\nSaved Excel to {output_path}")


def main():
    args = parse_args()
    fmt = args.format or ("json" if args.output.endswith(".json") else "xlsx")

    print(f"Cut Point Percentile Equivalents — method: {args.method}")
    print(f"Loading data from: {args.data_dir}")
    print(f"Cut points: {args.cut_points}\n")

    all_results, all_dist_context = load_and_calculate(
        args.data_dir, args.cut_points, args.method, args.years
    )

    if not all_results:
        print("ERROR: No data loaded. Check paths.")
        sys.exit(1)

    if fmt == "json":
        save_json(all_results, all_dist_context, args.output, args.method)
    else:
        save_xlsx(all_results, all_dist_context, args.output, args.method)


if __name__ == "__main__":
    main()
